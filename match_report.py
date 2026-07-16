import os
import openpyxl
import re
from datetime import datetime
from typing import Dict, List, Optional, Tuple

# ============================================================
# 0. 别名映射表（将用户常用简称映射到学校全称）
# ============================================================

SCHOOL_ALIAS = {
    "北大": "北京大学",
    "上大": "上海大学",
    "北邮": "北京邮电大学",
    "北航": "北京航空航天大",
    "北航大": "北京航空航天大",
    "川大": "四川大学",
    "中农": "中国农业大学",
    "中农大": "中国农业大学",
    "安农": "安徽农业大学",
    "安农大": "安徽农业大学",
    "清华": "清华大学",
    "西交": "西安交通大学",
    "西交大": "西安交通大学",
    "中大": "中山大学",
    "复旦": "复旦大学",
    "大工": "大连理工大学",
    "大工大": "大连理工大学",
    "对外经贸": "对外经济贸易大",
    "外经贸": "对外经济贸易大",
    "上电": "上海电力大学",
    "山财": "山东财经大学",
    "华科": "华中科技大学",
    "华中大": "华中科技大学",
    "同济": "同济大学",
    "北师大": "北京师范大学",
    "北师": "北京师范大学",
    "广大": "广州大学",
    "西工大": "西北工业大学",
    "西电": "西安电子科技大",
    "华工": "华南理工大学",
    "华南理工": "华南理工大学",
    "上工程": "上海工程技术大",
    "工程大": "上海工程技术大",
    "南宁二中": "南宁市第二中学",
    "哈医大": "哈尔滨医科大学",
    "福医大": "福建医科大学",
    "南开": "南开大学",
    "中央民大": "中央民族大学",
    "民大": "中央民族大学",
    "华南师大": "华南师范大学",
    "华师": "华南师范大学",
    "郑大": "郑州大学",
    "宁诺": "宁波诺丁汉大学",
    "华农": "华南农业大学",
    "福大": "福州大学",
    "东大": "东北大学",   # 与东南大学冲突，但匹配优先级会处理
    "中南财大": "中南财经政法大",
    "中南财经": "中南财经政法大",
    "二工大": "上海第二工业大",
    "上二工大": "上海第二工业大",
    "北交": "北京交通大学",
    "北交大": "北京交通大学",
    "上交": "上海交通大学",
    "上交大": "上海交通大学",
    "华东理工": "华东理工大学",
    "华理": "华东理工大学",
    "北理": "北京理工大学",
    "北理工": "北京理工大学",
    "河海": "河海大学",
    "西大": "广西大学",
    "武大": "武汉大学",
    "建桥": "上海建桥学院",
    "建桥学院": "上海建桥学院",
    "汕大": "汕头大学",
    "大外": "大连外国语大学",
    "西农": "西北农林科技大",
    "西北农大": "西北农林科技大",
    "中南": "中南大学",
    "东农": "东北农业大学",
    "东北农大": "东北农业大学",
    "哈工大": "哈尔滨工业大学",
    "天大": "天津大学",
    "江大": "江南大学",
    "华东师大": "华东师范大学",
    "吉大": "吉林大学",
    "上财": "上海财经大学",
    "南航": "南京航空航天大",
    "上师": "上海师范大学",
    "上师大": "上海师范大学",
    "南科": "南方科技大学",
    "南科大": "南方科技大学",
    "宁大": "宁波大学",
    "东林": "东北林业大学",
    "太理": "太原理工大学",
    "太原理工": "太原理工大学",
    "南大": "南京大学",
    "北信": "北京信息科技大",
    "北信科": "北京信息科技大",
}

def resolve_school_alias(keyword: str) -> str:
    """处理别名映射，并将以“大”结尾的关键词自动补全为“大学”。"""
    keyword = keyword.strip()
    if keyword in SCHOOL_ALIAS:
        return SCHOOL_ALIAS[keyword]
    if keyword.endswith("大"):
        return keyword[:-1] + "大学"
    return keyword

# ============================================================
# 1. 核心数据读取函数（含精确匹配及分数返回）
# ============================================================

def find_school_row(sheet, school_keyword: str, col_index: int = 1, start_row: int = 5) -> Tuple[Optional[int], Optional[str], int]:
    """返回 (行号, 学校全名, 匹配分数) 分数越高越精确。"""
    school_keyword = school_keyword.strip()
    if not school_keyword:
        return None, None, 0

    best_row = None
    best_name = None
    best_score = -1

    for row_idx in range(start_row, sheet.max_row + 1):
        cell_value = sheet.cell(row=row_idx, column=col_index).value
        if not cell_value:
            continue
        cell_str = str(cell_value).strip()

        if school_keyword in cell_str:
            score = 0
            if cell_str.startswith(school_keyword):
                score = 3
            elif re.search(r'(^|[\s,，、;；:：()（）])' + re.escape(school_keyword) + r'(?=[\s,，、;；:：()（）]|$)', cell_str):
                score = 2
            else:
                score = 1

            if score > best_score:
                best_score = score
                best_row = row_idx
                best_name = cell_str
            elif score == best_score and best_row is not None:
                if len(cell_str) < len(best_name):
                    best_row = row_idx
                    best_name = cell_str

    if best_row is not None:
        return best_row, best_name, best_score
    return None, None, 0

def parse_rank_from_detail(detail_value) -> int:
    if not detail_value or not isinstance(detail_value, str):
        return 0
    match = re.search(r'#(\d)', detail_value)
    if match:
        return int(match.group(1))
    return 0

def split_player_entries(raw_list: List[str]) -> List[str]:
    entries = []
    for item in raw_list:
        if re.match(r'^\s*#\d+\s*$', item):
            continue
        if '#' in item:
            parts = item.split('|')
            for part in parts:
                part = part.strip()
                if part and '#' in part:
                    entries.append(part)
    return entries

def load_school_data(file_path: str, school_keyword: str) -> Dict:
    wb = openpyxl.load_workbook(file_path, data_only=True)

    data = {
        "school": "",
        "east_score": 0.0,
        "east_rank": 0,
        "south_score": 0.0,
        "south_rank": 0,
        "total_score": 0.0,
        "final_rank": 0,
        "east_detail_raw": [],
        "south_detail_raw": [],
        "east_rank_1": 0,
        "east_rank_2": 0,
        "east_rank_3": 0,
        "east_rank_4": 0,
        "south_rank_1": 0,
        "south_rank_2": 0,
        "south_rank_3": 0,
        "south_rank_4": 0,
        "match_score": 0,
    }

    merge_sheet = wb["合并排行榜"]
    row_idx, full_name, score = find_school_row(merge_sheet, school_keyword, col_index=2, start_row=7)
    if row_idx:
        data["school"] = full_name
        data["match_score"] = score
        data["final_rank"] = merge_sheet.cell(row=row_idx, column=1).value or 0
        data["total_score"] = merge_sheet.cell(row=row_idx, column=9).value or 0.0

    east_sheet = wb["东风赛道"]
    row_idx, full_name, _ = find_school_row(east_sheet, school_keyword, col_index=2, start_row=5)
    if row_idx:
        data["east_rank"] = east_sheet.cell(row=row_idx, column=1).value or 0
        data["east_score"] = east_sheet.cell(row=row_idx, column=3).value or 0.0
        raw = []
        for c in range(4, east_sheet.max_column + 1):
            val = east_sheet.cell(row=row_idx, column=c).value
            if val is not None:
                raw.append(str(val).strip())
        data["east_detail_raw"] = raw

        entries = split_player_entries(raw)
        r1 = r2 = r3 = r4 = 0
        for entry in entries:
            rank = parse_rank_from_detail(entry)
            if rank == 1:
                r1 += 1
            elif rank == 2:
                r2 += 1
            elif rank == 3:
                r3 += 1
            elif rank == 4:
                r4 += 1
        data["east_rank_1"], data["east_rank_2"], data["east_rank_3"], data["east_rank_4"] = r1, r2, r3, r4

    south_sheet = wb["半庄赛道"]
    row_idx, full_name, _ = find_school_row(south_sheet, school_keyword, col_index=2, start_row=5)
    if row_idx:
        data["south_rank"] = south_sheet.cell(row=row_idx, column=1).value or 0
        data["south_score"] = south_sheet.cell(row=row_idx, column=3).value or 0.0
        raw = []
        for c in range(4, south_sheet.max_column + 1):
            val = south_sheet.cell(row=row_idx, column=c).value
            if val is not None:
                raw.append(str(val).strip())
        data["south_detail_raw"] = raw

        entries = split_player_entries(raw)
        r1 = r2 = r3 = r4 = 0
        for entry in entries:
            rank = parse_rank_from_detail(entry)
            if rank == 1:
                r1 += 1
            elif rank == 2:
                r2 += 1
            elif rank == 3:
                r3 += 1
            elif rank == 4:
                r4 += 1
        data["south_rank_1"], data["south_rank_2"], data["south_rank_3"], data["south_rank_4"] = r1, r2, r3, r4

    wb.close()
    return data

# ============================================================
# 2. 明细格式化（过滤冗余标记）
# ============================================================

def format_details(raw_list: List[str]) -> str:
    if not raw_list:
        return "暂无详细数据"

    entries = split_player_entries(raw_list)
    entries = [e for e in entries if not re.match(r'^\s*#\d+\s*$', e)]
    if not entries:
        return "暂无选手明细"

    current_scores = []
    result_lines = []

    for item in raw_list:
        if re.match(r'^\s*#\d+\s*$', item):
            continue
        if '|' in item and '#' not in item:
            current_scores = [s.strip() for s in item.split('|') if s.strip()]
            continue
        if '#' in item:
            player_parts = [p.strip() for p in item.split('|') if p.strip()]
            player_parts = [p for p in player_parts if not re.match(r'^\s*#\d+\s*$', p)]
            if not player_parts:
                continue
            if len(current_scores) == len(player_parts):
                for i, entry in enumerate(player_parts):
                    score = current_scores[i] if i < len(current_scores) else "?"
                    match = re.search(r'(.+?)\s*#(\d)\s*(\d+)', entry)
                    if match:
                        name = match.group(1).strip()
                        rank = match.group(2)
                        points = match.group(3)
                        result_lines.append(f"  {rank}位 {name} 点数{points} 得分{score}")
                    else:
                        result_lines.append(f"  {entry} 得分{score}")
            else:
                for entry in player_parts:
                    match = re.search(r'(.+?)\s*#(\d)\s*(\d+)', entry)
                    if match:
                        name = match.group(1).strip()
                        rank = match.group(2)
                        points = match.group(3)
                        result_lines.append(f"  {rank}位 {name} 点数{points}")
                    else:
                        result_lines.append(f"  {entry}")
            current_scores = []
            continue

        if '#' in item:
            match = re.search(r'(.+?)\s*#(\d)\s*(\d+)', item)
            if match:
                name = match.group(1).strip()
                rank = match.group(2)
                points = match.group(3)
                score = current_scores[0] if current_scores else "?"
                result_lines.append(f"  {rank}位 {name} 点数{points} 得分{score}")
                if current_scores:
                    current_scores = current_scores[1:]
            else:
                result_lines.append(f"  {item}")

    if not result_lines:
        return "暂无选手明细"
    return "\n".join(result_lines)

# ============================================================
# 3. 文件路径查找辅助
# ============================================================

def find_file(filename: str) -> str:
    if os.path.exists(filename):
        return os.path.abspath(filename)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    alt_path = os.path.join(script_dir, filename)
    if os.path.exists(alt_path):
        return alt_path
    cwd_path = os.path.join(os.getcwd(), filename)
    if os.path.exists(cwd_path):
        return cwd_path
    return filename

# ============================================================
# 4. 战报生成函数（含别名映射和匹配精度提醒）
# ============================================================

def generate_weekly_report(
    current_week_file: str,
    school_keyword: str = "第二工业",
    last_week_file: Optional[str] = None,
    week_number: Optional[int] = None
) -> str:
    # 别名映射与自动补全（不输出日志）
    original_keyword = school_keyword
    school_keyword = resolve_school_alias(school_keyword)

    current_week_file = find_file(current_week_file)
    if not os.path.exists(current_week_file):
        return f"❌ 错误：找不到文件 '{current_week_file}'"

    try:
        this_week = load_school_data(current_week_file, school_keyword)
    except Exception as e:
        return f"❌ 读取文件失败：{str(e)}"

    if not this_week['school']:
        return f"❌ 未找到学校：'{original_keyword}' 未参加第五届联合杯，或学校名称不匹配。"

    school_name = this_week['school']
    if school_name == "上海第二工业大":
        school_name = "上海第二工业大学"

    match_score = this_week.get('match_score', 0)
    precision_hint = ""
    if match_score < 3:
        precision_hint = "\n⚠️ 匹配精度较低，如需精确查询，请使用学校全称。"

    last_week = None
    if last_week_file:
        last_week_file = find_file(last_week_file)
        if os.path.exists(last_week_file):
            try:
                last_week = load_school_data(last_week_file, school_keyword)
                if not last_week.get('school'):
                    last_week = None
            except:
                last_week = None

    if last_week:
        delta_total = this_week["total_score"] - last_week["total_score"]
        delta_east = this_week["east_score"] - last_week["east_score"]
        delta_south = this_week["south_score"] - last_week["south_score"]
        last_total = last_week["total_score"]
        rank_change = last_week["final_rank"] - this_week["final_rank"]
    else:
        delta_total = this_week["total_score"]
        delta_east = this_week["east_score"]
        delta_south = this_week["south_score"]
        last_total = 0.0
        rank_change = 0

    if week_number is None:
        match = re.search(r'第(\d+)周', current_week_file)
        if match:
            week_number = int(match.group(1))
            week_title = f"第{week_number}周战报"
        else:
            week_title = "本周战报"
    else:
        week_title = f"第{week_number}周战报"

    rank_desc = f"↑ {rank_change}" if rank_change > 0 else f"↓ {-rank_change}" if rank_change < 0 else "持平"

    total_1 = this_week["east_rank_1"] + this_week["south_rank_1"]
    total_2 = this_week["east_rank_2"] + this_week["south_rank_2"]
    total_3 = this_week["east_rank_3"] + this_week["south_rank_3"]
    total_4 = this_week["east_rank_4"] + this_week["south_rank_4"]
    rank_dist = f"1位{total_1}次 / 2位{total_2}次 / 3位{total_3}次 / 4位{total_4}次"

    report = f"""
📊 第五届联合杯 · {week_title}

🏫 参赛学校：{school_name}
📅 报告生成：{datetime.now().strftime('%Y-%m-%d %H:%M')}

─────────────────────────────
📈 本周战况：
• 上周累计分数：{last_total:.1f} 分
• 本周新增分数：{delta_total:+.1f} 分
• 当前累计分数：{this_week['total_score']:.1f} 分
• 当前排名：第 {this_week['final_rank']} 名（{rank_desc}）
• 本周顺位分布：{rank_dist}

🀀 东风赛道：{this_week['east_score']:.1f} 分（{delta_east:+.1f} 分，排名 {this_week['east_rank']}）
  顺位：1位{this_week['east_rank_1']}次 / 2位{this_week['east_rank_2']}次 / 3位{this_week['east_rank_3']}次 / 4位{this_week['east_rank_4']}次

🀁 半庄赛道：{this_week['south_score']:.1f} 分（{delta_south:+.1f} 分，排名 {this_week['south_rank']}）
  顺位：1位{this_week['south_rank_1']}次 / 2位{this_week['south_rank_2']}次 / 3位{this_week['south_rank_3']}次 / 4位{this_week['south_rank_4']}次

─────────────────────────────
📌 本周明细（东风）：
{format_details(this_week['east_detail_raw'])}
📌 本周明细（半庄）：
{format_details(this_week['south_detail_raw'])}

─────────────────────────────
🎯 下周对阵：待定（功能开发中）
{precision_hint}
"""
    return report

# ============================================================
# 5. 测试入口 & 为 Agent 提供别名
# ============================================================

if __name__ == "__main__":
    test_file = "第五届联合杯_双赛道合并排行榜0712.xlsx"
    report = generate_weekly_report(
        current_week_file=test_file,
        school_keyword="第二工业",
        last_week_file=None,
        week_number=None
    )
    print(report)

# 为 Agent 调用提供别名
generate_weekly_report_text = generate_weekly_report